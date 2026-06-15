#!/usr/bin/env python3
"""Generate Source/ASM/level_runtime_table.inc from the audited HD-ref XLSX.

This table keeps metadata for all 22 adventure boards. Only explicitly listed
compiled_asset_slots get unique SPG pages; the rest fall back to level 1 until
a future level-pack/runtime loader or per-level overlay build exists.
"""
from __future__ import annotations

from decimal import Decimal
import math
import re
import struct
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent.parent
REF_LEVELS = Path.home() / "Desktop" / "Zuma-Deluxe-HD-ref" / "content" / "levels"
XLSX = REF_LEVELS / "zuma_levels_parameters.xlsx"
OUT = ROOT / "Source" / "ASM" / "level_runtime_table.inc"
CHAIN_OUT = ROOT / "Source" / "ASM" / "level_chain_table.inc"
CHAIN_BIN_OUT = ROOT / "Build" / "level_chain_table.bin"
LEVELS_HD = Path.home() / "Desktop" / "Zuma Deluxe" / "graphics" / "levels-HD"
LOCAL_LEVELS = ROOT / "Graphics" / "levels"
ADVENTURE_XLSX = LOCAL_LEVELS / "zuma_levels_parameters_adventure.xlsx"
GAUNTLET_XLSX = LOCAL_LEVELS / "gauntlet_level_settings.xlsx"

COMPILED_ASSET_SLOTS = {
    "spiral": ("BG_L01_FIRST_PAGE", "BG_L01_PALETTE_PAGE", "TRACK_L01_PAGE"),
}

SETTING_FIELDS = (
    "speed",
    "start",
    "score",
    "colors",
    "repeat",
    "single",
    "slowfactor",
    "partime",
)


def i(value, default=0) -> int:
    if value in (None, ""):
        return default
    return int(Decimal(str(value).strip()))


def scaled(value, mul: int, default=0) -> int:
    if value in (None, ""):
        return default
    return int((Decimal(str(value).strip()) * mul).to_integral_value())


def load_settings(wb):
    ws = wb["Difficulty settings"]
    headers = [c.value for c in ws[1]]
    rows = []
    index = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        data = dict(zip(headers, row))
        lvl_id = str(data["lvl_id"]).strip()
        index[lvl_id] = len(rows)
        rows.append(data)
    return rows, index


def load_graphics(wb):
    ws = wb["Boards Graphics"]
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        data = dict(zip(headers, row))
        out[str(data["Graphic ID"]).strip()] = data
    return out


def split_list(value) -> list[str]:
    if value in (None, ""):
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def load_adventure_chain(settings_index: dict[str, int], board_index: dict[str, int]) -> list[tuple[int, int]]:
    wb = openpyxl.load_workbook(ADVENTURE_XLSX, data_only=True)
    ws = wb["StageProgression"]
    headers = [c.value for c in ws[1]]
    chain: list[tuple[int, int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        boards = split_list(data.get("Boards"))
        lvl_ids = split_list(data.get("Lvl IDs"))
        for gid, lvl_id in zip(boards, lvl_ids):
            chain.append((board_index[gid], settings_index[lvl_id]))
    return chain


def load_gauntlet_table(settings_index: dict[str, int]) -> list[list[int]]:
    wb = openpyxl.load_workbook(GAUNTLET_XLSX, data_only=True)
    ws = wb["Gauntlet settings"]
    headers = [c.value for c in ws[1]]
    id_columns = ("Rabbit ID", "Eagle ID", "Jaguar ID", "Sun God ID")
    rows: list[list[int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("Graphic ID"):
            continue
        row_ids: list[int] = []
        for col in id_columns:
            lvl_id = data.get(col)
            row_ids.append(settings_index.get(str(lvl_id).strip(), 0xFF) if lvl_id else 0xFF)
        rows.append(row_ids)
    return rows


def parse_dat(path: Path):
    data = path.read_bytes()
    o = 0x10
    count1 = struct.unpack_from("<i", data, o)[0]
    o += 4 + count1 * 10
    curve_len = struct.unpack_from("<i", data, o)[0]
    o += 4
    x = struct.unpack_from("<f", data, o)[0]
    o += 4
    y = struct.unpack_from("<f", data, o)[0]
    o += 4
    pts = [(x, y)]
    for _ in range(curve_len):
        if o + 4 > len(data):
            break
        dx = struct.unpack_from("<b", data, o + 2)[0]
        dy = struct.unpack_from("<b", data, o + 3)[0]
        o += 4
        x += dx / 100.0
        y += dy / 100.0
        pts.append((x, y))
    return pts


def resampled_last_point(points, step=1.0):
    cum = [0.0]
    for idx in range(1, len(points)):
        dx = points[idx][0] - points[idx - 1][0]
        dy = points[idx][1] - points[idx - 1][1]
        cum.append(cum[-1] + math.hypot(dx, dy))
    total_len = cum[-1]
    src_i = 0
    last = points[0]
    for sample in range(int(total_len / step) + 1):
        target = sample * step
        while src_i < len(points) - 2 and cum[src_i + 1] < target:
            src_i += 1
        seg = cum[src_i + 1] - cum[src_i]
        t = (target - cum[src_i]) / seg if seg > 0 else 0.0
        x = points[src_i][0] + t * (points[src_i + 1][0] - points[src_i][0])
        y = points[src_i][1] + t * (points[src_i + 1][1] - points[src_i][1])
        last = (int(round(x)), int(round(y)))
    return last


def killzone_center_for(gid: str) -> tuple[int, int]:
    path = LEVELS_HD / gid / f"{gid}.dat"
    if not path.exists():
        matches = [p for p in LEVELS_HD.iterdir() if p.is_dir() and p.name.lower() == gid.lower()]
        if matches:
            path = matches[0] / f"{matches[0].name}.dat"
            if not path.exists():
                dats = sorted(matches[0].glob("*.dat"))
                if dats:
                    path = dats[0]
    if not path.exists():
        return 0, 0
    return resampled_last_point(parse_dat(path))


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    settings, settings_index = load_settings(wb)
    graphics = load_graphics(wb)
    ws = wb["Boards x Tiers"]
    headers = [c.value for c in ws[1]]

    board_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        data = dict(zip(headers, row))
        gid = str(data["Graphic ID"]).strip()
        g = graphics[gid]
        tier_ids = []
        for tier in range(1, 5):
            lvl_id = data.get(f"T{tier}_lvl")
            tier_ids.append(settings_index.get(str(lvl_id).strip(), 0xFF) if lvl_id else 0xFF)
        kx, ky = killzone_center_for(gid)
        board_rows.append((gid, str(data["Display Name"]).strip(), tier_ids, i(g.get("FrogX (gx)")), i(g.get("FrogY (gy)")), kx, ky))

    board_index = {gid: idx for idx, (gid, *_rest) in enumerate(board_rows)}
    adventure_chain = load_adventure_chain(settings_index, board_index)
    gauntlet_table = load_gauntlet_table(settings_index)

    lines = [
        "; Auto-generated by Source/OTHER/make_level_runtime_table.py.",
        "; Source: Zuma-Deluxe-HD-ref/content/levels/zuma_levels_parameters.xlsx",
        ";",
        "; 22-board metadata table is resident and cheap.",
        "; Only compiled asset slots below allocate unique SPG pages.",
        "",
        "LEVEL_RT_RECORD_SIZE EQU 15",
        f"LEVEL_RUNTIME_COUNT  EQU {len(board_rows)}",
        "LEVEL_SETTING_COUNT  EQU {0}".format(len(settings)),
        "ADVENTURE_CHAIN_COUNT EQU {0}".format(len(adventure_chain)),
        "GAUNTLET_SETTING_COUNT EQU {0}".format(len(gauntlet_table) * 4),
        "CHAIN_TABLE_PAGE    EQU #42",
        "ADVENTURE_CHAIN_OFFSET EQU #0000",
        "GAUNTLET_SETTING_OFFSET EQU ADVENTURE_CHAIN_COUNT * 2",
        "LEVEL_RT_BG_PAGE     EQU 0",
        "LEVEL_RT_PAL_PAGE    EQU 1",
        "LEVEL_RT_TRACK_PAGE  EQU 2",
        "LEVEL_RT_TIER1       EQU 3",
        "LEVEL_RT_TIER2       EQU 4",
        "LEVEL_RT_TIER3       EQU 5",
        "LEVEL_RT_TIER4       EQU 6",
        "LEVEL_RT_FROG_X      EQU 7",
        "LEVEL_RT_FROG_Y      EQU 9",
        "LEVEL_RT_KZ_X        EQU 11",
        "LEVEL_RT_KZ_Y        EQU 13",
        "",
        "BG_L01_FIRST_PAGE    EQU #07",
        "BG_L01_PALETTE_PAGE  EQU #11",
        "TRACK_L01_PAGE       EQU #06",
        "",
        "; Current compiled fallback backend: level 1 only.",
        "; Runtime loads the selected board graphics/track from ZUMALVL.PAK.",
        "LevelRuntimeTable:",
    ]

    for idx, (gid, disp, tiers, fx, fy, kx, ky) in enumerate(board_rows):
        bg_page, pal_page, track_page = COMPILED_ASSET_SLOTS.get(gid, COMPILED_ASSET_SLOTS["spiral"])
        lines.append(f"                ; {idx + 1:02d}: {gid} / {disp}")
        lines.append(
            f"                DB {bg_page}, {pal_page}, {track_page}, "
            + ", ".join(f"#{x:02X}" if x == 0xFF else str(x) for x in tiers)
        )
        # 1024×768 порт: frog/killzone позиции ×1.6 (640→1024), как трек/шары.
        sx = lambda v: int(round(v * 8 / 5))
        lines.append(f"                DW {sx(fx)}, {sx(fy)}, {sx(kx)}, {sx(ky)}")

    lines += [
        "",
        "LevelSettingsTable:",
        "                ; speed_x100, start, score word, colors, repeat, single, slowfactor_x10, partime",
    ]

    for idx, row in enumerate(settings):
        lvl_id = str(row["lvl_id"]).strip()
        speed = scaled(row.get("speed"), 100)
        start = i(row.get("start"))
        score = i(row.get("score"))
        colors = i(row.get("colors"), 6)
        repeat = i(row.get("repeat"), 50)
        single = i(row.get("single"), 0)
        slowfactor = scaled(row.get("slowfactor"), 10)
        partime = i(row.get("partime"))
        lines.append(f"                ; {idx:02d}: {lvl_id}")
        lines.append(f"                DB {speed}, {start}")
        lines.append(f"                DW {score}")
        lines.append(f"                DB {colors}, {repeat}, {single}, {slowfactor}, {partime}")

    chain_lines = [
        "; Auto-generated by Source/OTHER/make_level_runtime_table.py.",
        "; Source: Graphics/levels/zuma_levels_parameters_adventure.xlsx and gauntlet_level_settings.xlsx",
        ";",
        "; Держим отдельно от level_runtime_table.inc: тот include живёт в LOG_BLOCK,",
        "; а эти цепочки должны лежать в свободном resident Core/Main0.",
        "",
        "",
        "AdventureChainTable:",
        "                ; пары: CurrentLevel, CurrentSettingIndex",
    ]
    for idx, (board_id, setting_id) in enumerate(adventure_chain):
        gid = board_rows[board_id][0]
        chain_lines.append(f"                DB {board_id}, {setting_id}                 ; {idx + 1:02d}: {gid}")

    chain_lines += [
        "",
        "GauntletSettingTable:",
        "                ; 22 карты x 4 ранга: Rabbit, Eagle, Jaguar, Sun God",
    ]
    for idx, row_ids in enumerate(gauntlet_table):
        gid = board_rows[idx][0] if idx < len(board_rows) else f"row{idx + 1}"
        values = ", ".join(f"#{x:02X}" if x == 0xFF else str(x) for x in row_ids)
        chain_lines.append(f"                DB {values}                 ; {idx + 1:02d}: {gid}")

    chain_bin = bytearray()
    for board_id, setting_id in adventure_chain:
        chain_bin.extend((board_id & 0xFF, setting_id & 0xFF))
    for row_ids in gauntlet_table:
        chain_bin.extend(value & 0xFF for value in row_ids)

    lines += [
        "",
        "LevelSelectApplyLevelClick:",
        "                LD   A, (Core.LevelSelectNextClick)",
        "                OR   A",
        "                JR   Z, .back",
        "                LD   A, (CurrentLevel)",
        "                INC  A",
        "                CP   LEVEL_RUNTIME_COUNT",
        "                JR   C, .store",
        "                XOR  A",
        ".store:         LD   (CurrentLevel), A",
        "                CALL Core.LoadLevelSelectPreviewAssets",
        "                RET",
        ".back:          LD   A, (Core.LevelSelectBackClick)",
        "                OR   A",
        "                RET  Z",
        "                LD   A, (CurrentLevel)",
        "                OR   A",
        "                JR   Z, .wrap_last",
        "                DEC  A",
        "                JR   .store",
        ".wrap_last:     LD   A, LEVEL_RUNTIME_COUNT - 1",
        "                JR   .store",
        "",
    ]

    OUT.write_text("\n".join(lines), encoding="utf-8")
    CHAIN_OUT.write_text("\n".join(chain_lines), encoding="utf-8")
    CHAIN_BIN_OUT.parent.mkdir(parents=True, exist_ok=True)
    CHAIN_BIN_OUT.write_bytes(bytes(chain_bin))
    print(f"wrote {OUT} ({len(board_rows)} boards, {len(settings)} settings)")
    print(f"wrote {CHAIN_OUT} ({len(adventure_chain)} adventure pairs, {len(gauntlet_table) * 4} gauntlet entries)")
    print(f"wrote {CHAIN_BIN_OUT} ({len(chain_bin)} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
