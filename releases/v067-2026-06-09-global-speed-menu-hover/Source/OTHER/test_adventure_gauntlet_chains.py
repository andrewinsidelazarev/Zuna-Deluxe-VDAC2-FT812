#!/usr/bin/env python3
"""Проверка раздельных цепочек Adventure/Gauntlet без запуска игры."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from zuma_full_z80_emulator import ZumaFullZ80Emulator


ROOT = Path(__file__).resolve().parents[2]
ASM = ROOT / "Source" / "ASM"
LEVELS = ROOT / "Graphics" / "levels"


def split_list(value) -> list[str]:
    if value in (None, ""):
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def setting_index() -> dict[str, int]:
    wb = openpyxl.load_workbook(LEVELS / "zuma_levels_parameters_adventure.xlsx", data_only=True)
    ws = wb["Difficulty settings"]
    out: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            out[str(row[0]).strip()] = len(out)
    return out


def board_index() -> dict[str, int]:
    wb = openpyxl.load_workbook(LEVELS / "zuma_levels_parameters_adventure.xlsx", data_only=True)
    ws = wb["Boards x Tiers"]
    out: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            out[str(row[1]).strip()] = len(out)
    return out


def expected_adventure(si: dict[str, int], bi: dict[str, int]) -> list[tuple[int, int]]:
    wb = openpyxl.load_workbook(LEVELS / "zuma_levels_parameters_adventure.xlsx", data_only=True)
    ws = wb["StageProgression"]
    rows: list[tuple[int, int]] = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        for gid, lvl_id in zip(split_list(data.get("Boards")), split_list(data.get("Lvl IDs"))):
            rows.append((bi[gid], si[lvl_id]))
    return rows


def expected_gauntlet(si: dict[str, int]) -> list[list[int]]:
    wb = openpyxl.load_workbook(LEVELS / "gauntlet_level_settings.xlsx", data_only=True)
    ws = wb["Gauntlet settings"]
    headers = [cell.value for cell in ws[1]]
    cols = ("Rabbit ID", "Eagle ID", "Jaguar ID", "Sun God ID")
    rows: list[list[int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("Graphic ID"):
            continue
        rows.append([si.get(str(data[col]).strip(), 0xFF) if data.get(col) else 0xFF for col in cols])
    return rows


def db_rows(label: str) -> list[list[int]]:
    text = (ASM / "level_chain_table.inc").read_text(encoding="utf-8")
    tail = text.split(label + ":", 1)[1]
    tail = tail.split("\n\n", 1)[0]
    rows: list[list[int]] = []
    for line in tail.splitlines():
        line = line.split(";", 1)[0]
        if "DB" not in line:
            continue
        nums = []
        for token in line.split("DB", 1)[1].split(","):
            token = token.strip()
            if not token:
                continue
            nums.append(int(token[1:], 16) if token.startswith("#") else int(token))
        rows.append(nums)
    return rows


def expected_chain_blob(adv: list[tuple[int, int]], gaunt: list[list[int]]) -> bytes:
    out = bytearray()
    for board_id, setting_id in adv:
        out.extend((board_id & 0xFF, setting_id & 0xFF))
    for row in gaunt:
        out.extend(value & 0xFF for value in row)
    return bytes(out)


def main() -> int:
    si = setting_index()
    bi = board_index()
    adv = expected_adventure(si, bi)
    gaunt = expected_gauntlet(si)

    assert [tuple(row) for row in db_rows("AdventureChainTable")] == adv
    assert db_rows("GauntletSettingTable") == gaunt
    assert (ROOT / "Build" / "level_chain_table.bin").read_bytes() == expected_chain_blob(adv, gaunt)
    assert len(adv) == 76
    assert len(gaunt) == 22 and all(len(row) == 4 for row in gaunt)
    assert adv[0] == (0, 0)
    assert adv[-1] == (21, 75)
    assert gaunt[6][1] == si["lvl66"]  # Long Range Eagle — заполненный финальный Eagle-пресет

    runtime = (ASM / "level_runtime_table.inc").read_text(encoding="utf-8")
    assert "AdventureChainTable:" not in runtime
    assert "GauntletSettingTable:" not in runtime

    menu = (ASM / "MenuMain.asm").read_text(encoding="utf-8")
    assert "JP   FadeMenuToAdventure" in menu
    assert "JP   FadeMenuToLevelSelect" in menu

    main_asm = (ASM / "main.asm").read_text(encoding="utf-8")
    spg = (ROOT / "spgbld_vdac2.ini").read_text(encoding="utf-8")
    assert "Block = #0000, #42, Build/level_chain_table.bin" in spg
    assert 'include "level_chain_table.inc"' not in main_asm
    assert "CALL OVL_ResolveCurrentModeSelection" in main_asm

    emu = ZumaFullZ80Emulator(ROOT)
    sym = emu.sym
    emu.mem.pages[3] = 0x40

    emu.set_byte(sym["Core.AdventurePos"], 75)
    emu.set_byte(sym["Core.CurrentGameMode"], 0)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 21
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == 75

    emu.set_byte(sym["Core.AdventurePos"], 99)
    emu.set_byte(sym["Core.CurrentGameMode"], 0)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.AdventurePos"]) == 0
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 0
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == 0

    emu.set_byte(sym["Core.CurrentLevel"], 6)
    emu.set_byte(sym["Core.CurrentDifficulty"], 1)
    emu.set_byte(sym["Core.CurrentGameMode"], 1)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == si["lvl66"]

    emu.set_byte(sym["Core.CurrentGameMode"], 0)
    emu.set_byte(sym["Core.AdventurePos"], 1)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 1
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == 1

    emu.set_byte(sym["Core.CurrentGameMode"], 1)
    emu.set_byte(sym["Core.CurrentLevel"], 13)
    emu.set_byte(sym["Core.CurrentDifficulty"], 1)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == si["lvl66"]

    print("PASS: Adventure/Gauntlet цепочки раздельные, таблицы совпадают с XLSX")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
