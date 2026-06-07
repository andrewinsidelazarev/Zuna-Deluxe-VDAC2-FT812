#!/usr/bin/env python3
"""build_levels_table.py — генерирует zuma_levels_parameters.xlsx по
content/levels/levels.xml + wiki-подобные источники.

Структура XLSX:
- Sheet "Boards x Tiers" — 22 board × 4 tier (+ space tier 13). Колонки:
  graphic, dispname, и для каждого tier (1..4): lvl_id, speed, start, score, colors,
  repeat, single, slowfactor, partime.
- Sheet "Difficulty settings" — все lvlXX settings из levels.xml в одну таблицу.
- Sheet "StageProgression" — какой board в каком stage с каким lvl_id.
- Sheet "Wiki notes" — суммарные wiki-described mechanics (win condition,
  spawn gate, scoring formula, repeatChance %).
"""
import re
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r'C:\Users\Администратор\Desktop\Zuma-Deluxe-HD-ref\content\levels')
XML = ROOT / 'levels.xml'
OUT = ROOT / 'zuma_levels_parameters.xlsx'


def parse_xml():
    """Returns (graphics_dict, graphics_full, settings_dict, stage_progression).
    Graphics: id → dispname
    Graphics_full: id → dict (curve, image, image-top, dispname, gx, gy, treasure_pts, cutouts, ...)
    Settings: id → dict of params (speed, start, score, colors, repeat, single, slowfactor, partime)
    StageProgression: stage_num → (board_list, lvl_id_list)
    """
    # Parse manually because XML has comments and irregular formatting
    text = XML.read_text(encoding='utf-8', errors='replace')

    graphics = {}
    graphics_full = {}
    # match each Graphics block from opening tag to closing
    for m in re.finditer(r'<Graphics\s+([^>]*?)(?:/>|>(.*?)</Graphics>)', text, re.DOTALL):
        attrs_str = m.group(1)
        body = m.group(2) or ''
        attrs = dict(re.findall(r'([\w-]+)="([^"]*)"', attrs_str))
        gid = attrs.get('id', '?')
        graphics[gid] = attrs.get('dispname', '?')
        # Extract child elements
        treasure_pts = []
        for tm in re.finditer(r'<TreasurePoint\s+([^/]+?)/>', body):
            t_attrs = dict(re.findall(r'(\w+)="([^"]*)"', tm.group(1)))
            treasure_pts.append(t_attrs)
        cutouts = []
        for cm in re.finditer(r'<Cutout\s+([^/]+?)/>', body):
            c_attrs = dict(re.findall(r'([\w-]+)="([^"]*)"', cm.group(1)))
            cutouts.append(c_attrs)
        bg_alphas = []
        for ba in re.finditer(r'<BackgroundAlpha\s+([^/]+?)/>', body):
            ba_attrs = dict(re.findall(r'([\w-]+)="([^"]*)"', ba.group(1)))
            bg_alphas.append(ba_attrs)
        attrs['treasure_pts'] = treasure_pts
        attrs['cutouts'] = cutouts
        attrs['bg_alphas'] = bg_alphas
        graphics_full[gid] = attrs
    graphics['space'] = 'Space'  # no dispname in xml
    graphics_full.setdefault('space', {'id': 'space', 'curve': 'space', 'gx': '246', 'gy': '188', 'dispname': 'Space',
                                       'treasure_pts': [], 'cutouts': [], 'bg_alphas': []})

    settings = {}
    for m in re.finditer(r'<Settings\s+([^/]+?)/>', text, re.DOTALL):
        attrs = dict(re.findall(r'(\w+)="([^"]*)"', m.group(1)))
        if 'id' in attrs:
            settings[attrs['id']] = attrs

    # Parse StageProgression (one big block)
    stage_prog = {}
    sp_match = re.search(r'<StageProgression(.*?)/>', text, re.DOTALL)
    if sp_match:
        sp_text = sp_match.group(1)
        for n in range(1, 14):
            sm = re.search(rf'stage{n}\s*=\s*"([^"]+)"', sp_text)
            dm = re.search(rf'diffi{n}\s*=\s*"([^"]+)"', sp_text)
            if sm and dm:
                boards = [s.strip() for s in sm.group(1).split(',')]
                lvls = [s.strip() for s in dm.group(1).split(',')]
                stage_prog[n] = (boards, lvls)
    return graphics, graphics_full, settings, stage_prog


def get_param(settings, lvl_id, key, default=''):
    if lvl_id not in settings:
        return default
    return settings[lvl_id].get(key, default)


def build_xlsx():
    graphics, graphics_full, settings, stage_prog = parse_xml()
    wb = openpyxl.Workbook()

    # === Sheet 1: Boards × Tiers ===
    ws = wb.active
    ws.title = 'Boards x Tiers'

    # 22 уникальных board ids (исключая space — отдельная строка)
    # Tier mapping: stages 1-3 = tier 1, 4-6 = tier 2, 7-9 = tier 3, 10-12 = tier 4, 13 = space
    BOARDS_ORDER = [
        'spiral', 'claw', 'riverbed', 'targetglyph', 'blackswirley', 'turnaround',
        'longrange', 'tiltspiral', 'underover', 'warshak', 'loopy', 'snakepit',
        'groovefest', 'spaceinvaders', 'triangle', 'coaster', 'squaresville',
        'tunnellevel', 'serpents', 'overunder', 'inversespiral', 'space',
    ]

    TIER_STAGES = {
        1: [1, 2, 3],
        2: [4, 5, 6],
        3: [7, 8, 9],
        4: [10, 11, 12],
    }

    # Header
    headers = ['#', 'Graphic ID', 'Display Name']
    for tier in (1, 2, 3, 4):
        for col in ['lvl', 'speed', 'start', 'score', 'colors', 'repeat', 'single', 'slowfactor', 'partime']:
            headers.append(f'T{tier}_{col}')
    ws.append(headers)

    # Style header
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='305496')
    thin = Side(border_style='thin', color='000000')
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Find board's lvl_id per tier
    def find_lvl_for_board_in_tier(board, tier):
        if board == 'space':
            return 'lvl131' if tier == 4 else ''
        for stage in TIER_STAGES.get(tier, []):
            if stage in stage_prog:
                boards, lvls = stage_prog[stage]
                if board in boards:
                    # Use first matching lvl_id from this tier's stages
                    # Pick the lvl whose index in diffi corresponds to position in stage
                    idx_in_stage = boards.index(board)
                    if idx_in_stage < len(lvls):
                        return lvls[idx_in_stage]
                    return lvls[0]
        return ''

    row_idx = 2
    tier_fills = {
        1: PatternFill('solid', fgColor='D9E1F2'),
        2: PatternFill('solid', fgColor='FFF2CC'),
        3: PatternFill('solid', fgColor='FCE4D6'),
        4: PatternFill('solid', fgColor='E2EFDA'),
    }

    for i, board in enumerate(BOARDS_ORDER, 1):
        row = [i, board, graphics.get(board, '?')]
        for tier in (1, 2, 3, 4):
            lvl_id = find_lvl_for_board_in_tier(board, tier)
            row.append(lvl_id)
            row.append(get_param(settings, lvl_id, 'speed'))
            row.append(get_param(settings, lvl_id, 'start'))
            row.append(get_param(settings, lvl_id, 'score'))
            row.append(get_param(settings, lvl_id, 'colors'))
            row.append(get_param(settings, lvl_id, 'repeat'))
            row.append(get_param(settings, lvl_id, 'single'))
            row.append(get_param(settings, lvl_id, 'slowfactor'))
            row.append(get_param(settings, lvl_id, 'partime'))
        ws.append(row)
        # Apply tier fills (3 fixed columns + 9 per tier)
        for tier in (1, 2, 3, 4):
            start_col = 3 + 1 + (tier - 1) * 9  # после "#, Graphic, Dispname" = 3; tier 1 starts at col 4
            for off in range(9):
                ws.cell(row=row_idx, column=start_col + off).fill = tier_fills[tier]
                ws.cell(row=row_idx, column=start_col + off).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col in (1, 2, 3):
            ws.cell(row=row_idx, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        row_idx += 1

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13 if col > 3 else 18
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22

    # Freeze panes after header row + first 3 columns
    ws.freeze_panes = 'D2'

    # === Sheet 2: Difficulty settings (raw) ===
    ws2 = wb.create_sheet('Difficulty settings')
    ws2.append(['lvl_id', 'speed', 'start', 'score', 'colors', 'repeat', 'single', 'slowfactor', 'partime',
                'reloaddelay', 'mergespeed', 'firespeed', 'zumaslow', 'zumaback'])
    for col in range(1, 15):
        c = ws2.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    # All Adventure mode lvlXX
    lvl_ids = sorted([k for k in settings if k.startswith('lvl')], key=lambda s: (len(s), s))
    for lvl in lvl_ids:
        s = settings[lvl]
        ws2.append([
            lvl, s.get('speed', ''), s.get('start', ''), s.get('score', ''),
            s.get('colors', ''), s.get('repeat', ''), s.get('single', ''),
            s.get('slowfactor', ''), s.get('partime', ''),
            s.get('reloaddelay', ''), s.get('mergespeed', ''), s.get('firespeed', ''),
            s.get('zumaslow', ''), s.get('zumaback', ''),
        ])
    for col in range(1, 15):
        ws2.column_dimensions[get_column_letter(col)].width = 13
    ws2.freeze_panes = 'A2'

    # === Sheet 3: StageProgression ===
    ws3 = wb.create_sheet('StageProgression')
    ws3.append(['Stage', 'Tier', 'Boards', 'Lvl IDs', '# Boards', '# Difficulty'])
    for col in range(1, 7):
        c = ws3.cell(row=1, column=col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    tier_lookup = {1:1, 2:1, 3:1, 4:2, 5:2, 6:2, 7:3, 8:3, 9:3, 10:4, 11:4, 12:4, 13:5}
    for n in sorted(stage_prog):
        boards, lvls = stage_prog[n]
        ws3.append([n, tier_lookup.get(n, '?'), ', '.join(boards), ', '.join(lvls), len(boards), len(lvls)])
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 7
    ws3.column_dimensions['C'].width = 80
    ws3.column_dimensions['D'].width = 50
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 12
    for row in range(2, ws3.max_row + 1):
        ws3.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws3.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

    # === Sheet 4: Wiki notes ===
    ws4 = wb.create_sheet('Wiki notes')
    notes = [
        ['Параметр', 'Описание', 'Wiki / HD-ref ссылка'],
        ['speed', 'Скорость движения цепи (множитель базовой)', 'Speedrun.com guide, HD-ref Level.h:ballSpd'],
        ['start', 'Стартовое количество шаров в цепи (ballStartCount)', 'HD-ref Level.h:25'],
        ['score', 'Целевой счёт для заполнения Zuma bar (gaugeScore)', 'HD-ref Level.h:27, Statistics.c'],
        ['colors', 'Количество цветов шаров на уровне', 'HD-ref Level.h:30'],
        ['repeat', 'repeatChance % — вероятность повтора предыдущего цвета (создаёт same-color кластеры в цепи)', 'HD-ref Level.h:28; default 50%'],
        ['single', 'singleChance % — вероятность одиночного шара (cluster size = 1)', 'HD-ref Level.h:29'],
        ['slowfactor', 'Множитель для slow-mode (когда цепь почти у killzone замедляется)', 'PopCap manual, HD-ref Level.h:33'],
        ['partime', 'Par time в секундах (для бонуса при быстром прохождении)', 'GameFAQs FAQ'],
        ['reloaddelay', 'Задержка между выстрелами (0 = быстро)', 'HD-ref Bullets.c'],
        ['mergespeed', 'Скорость объединения цепи после destroy', 'HD-ref BallChain.c'],
        ['firespeed', 'Скорость bullet', 'HD-ref Bullets.c'],
        ['zumaslow / zumaback', 'Параметры zuma slow-mode (редко используются)', 'HD-ref Level.h, lvl35 пример'],
        [],
        ['Условие победы', 'Заполнить Zuma bar до gaugeScore → spawn gate OFF → дочистить остаток шаров',
         'Zuma Wiki/Fandom, GameFAQs FAQ, official Xbox PopCap manual'],
        ['Условие поражения', 'Tail of chain достигает skull (killzone) → game over',
         'Manual + visual feedback'],
        ['Scoring формула', 'points = count*10 + combo*100 + (100 + 10*(chain-5) если chain>=5 без combo) + gap_bonus',
         'HD-ref Statistics.c:37'],
        ['Combo', 'Same color match-3 подряд → +100 × combo_count',
         'Manual: "COMBO!"  "COMBO X2!"'],
        ['Chain bonus', '≥5 explosions подряд (cascade) без combo → +100+10×(chain−5)',
         'Manual + GameFAQs'],
        ['Gap bonus', 'Bullet прошёл через дыру в цепи без попадания → +500×(MAX_GAP−distance)/MAX_GAP, clamp ≥10, ×2 для consecutive',
         'HD-ref Statistics.c:78 — Statistics_AddBulletGap'],
        ['ZUMA! voice', 'При gauge full воспроизводится sample "ZUMA!"',
         'Manual + sound'],
        [],
        ['Tier структура', '22 unique board × 4 tier difficulty (stages 1-3 / 4-6 / 7-9 / 10-12) + Space (stage 13)',
         'levels.xml StageProgression'],
        ['Стартовая фаза', 'Fast phase спавнит ballStartCount шаров быстро, потом normal speed',
         'HD-ref BallChain.c fastModeBallsCountdown=32'],
        ['Refresh circle', 'После 1 round'],
    ]
    for r in notes:
        ws4.append(r)
    for col in range(1, 4):
        c = ws4.cell(row=1, column=col)
        c.font = hdr_font; c.fill = hdr_fill
    ws4.column_dimensions['A'].width = 22
    ws4.column_dimensions['B'].width = 70
    ws4.column_dimensions['C'].width = 50
    for row in range(1, ws4.max_row + 1):
        for col in range(1, 4):
            ws4.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    # === Sheet 5: Boards Graphics — assets per board ===
    ws5 = wb.create_sheet('Boards Graphics')
    ws5.append(['#', 'Graphic ID', 'Display Name', 'Curve', 'Curve2', 'Image', 'Image-top',
                'FrogX (gx)', 'FrogY (gy)', 'SkullRot', 'Treasure pts (x,y,dist)',
                '# Cutouts', 'Cutouts', '# BG Alphas'])
    for col in range(1, 15):
        c = ws5.cell(row=1, column=col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, board in enumerate(BOARDS_ORDER, 1):
        g = graphics_full.get(board, {})
        tp_str = '; '.join(
            f'({p.get("x","?")},{p.get("y","?")},dist={p.get("dist1") or p.get("dist2","?")})'
            for p in g.get('treasure_pts', [])
        )
        cutout_str = ', '.join(c.get('image', '?') for c in g.get('cutouts', []))
        ws5.append([
            i, board, g.get('dispname', '?'), g.get('curve', '?'),
            g.get('curve2', ''),
            g.get('image', '?'), g.get('image-top', ''),
            g.get('gx', '?'), g.get('gy', '?'), g.get('skullrot', ''),
            tp_str, len(g.get('cutouts', [])), cutout_str, len(g.get('bg_alphas', []))
        ])
        for col in range(1, 15):
            ws5.cell(row=i+1, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [4, 18, 22, 14, 14, 14, 14, 11, 11, 10, 60, 11, 30, 12]
    for col, w in enumerate(widths, 1):
        ws5.column_dimensions[get_column_letter(col)].width = w
    ws5.freeze_panes = 'D2'

    wb.save(OUT)
    print(f'saved {OUT}')
    print(f'  Sheets: Boards x Tiers, Difficulty settings, StageProgression, Wiki notes, Boards Graphics')
    print(f'  Boards: {len(BOARDS_ORDER)}')
    print(f'  Settings: {len(settings)}')
    print(f'  Stages: {len(stage_prog)}')


if __name__ == '__main__':
    build_xlsx()
