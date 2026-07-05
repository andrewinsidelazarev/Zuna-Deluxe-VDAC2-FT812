#!/usr/bin/env python3
"""Проверка раздельных цепочек Adventure/Gauntlet без запуска игры."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from zuma_full_z80_emulator import ZumaFullZ80Emulator


ROOT = Path(__file__).resolve().parents[2]
ASM = ROOT / "Source" / "ASM"
LEVELS = ROOT / "Graphics" / "levels"


def split_list(value) -> list[str]:
    if value in (None, ""):
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def setting_index() -> dict[str, int]:
    wb = openpyxl.load_workbook(LEVELS / "zuma_levels_parameters_adventure.xlsx", data_only=True)
    ws = wb["Difficulty settings"]
    out: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            out[str(row[0]).strip()] = len(out)
    return out


def board_index() -> dict[str, int]:
    wb = openpyxl.load_workbook(LEVELS / "zuma_levels_parameters_adventure.xlsx", data_only=True)
    ws = wb["Boards x Tiers"]
    out: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            out[str(row[1]).strip()] = len(out)
    return out


def expected_adventure(si: dict[str, int], bi: dict[str, int]) -> list[tuple[int, int]]:
    wb = openpyxl.load_workbook(LEVELS / "zuma_levels_parameters_adventure.xlsx", data_only=True)
    ws = wb["StageProgression"]
    rows: list[tuple[int, int]] = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        for gid, lvl_id in zip(split_list(data.get("Boards")), split_list(data.get("Lvl IDs"))):
            rows.append((bi[gid], si[lvl_id]))
    return rows


def expected_gauntlet(si: dict[str, int]) -> list[list[int]]:
    wb = openpyxl.load_workbook(LEVELS / "gauntlet_level_settings.xlsx", data_only=True)
    ws = wb["Gauntlet settings"]
    headers = [cell.value for cell in ws[1]]
    cols = ("Rabbit ID", "Eagle ID", "Jaguar ID", "Sun God ID")
    rows: list[list[int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("Graphic ID"):
            continue
        rows.append([si.get(str(data[col]).strip(), 0xFF) if data.get(col) else 0xFF for col in cols])
    return rows


def db_rows(label: str) -> list[list[int]]:
    text = (ASM / "level_chain_table.inc").read_text(encoding="utf-8")
    tail = text.split(label + ":", 1)[1]
    tail = tail.split("\n\n", 1)[0]
    rows: list[list[int]] = []
    for line in tail.splitlines():
        line = line.split(";", 1)[0]
        if "DB" not in line:
            continue
        nums = []
        for token in line.split("DB", 1)[1].split(","):
            token = token.strip()
            if not token:
                continue
            nums.append(int(token[1:], 16) if token.startswith("#") else int(token))
        rows.append(nums)
    return rows


def expected_chain_blob(adv: list[tuple[int, int]], gaunt: list[list[int]]) -> bytes:
    out = bytearray()
    for board_id, setting_id in adv:
        out.extend((board_id & 0xFF, setting_id & 0xFF))
    for row in gaunt:
        out.extend(value & 0xFF for value in row)
    return bytes(out)


def patch_ret(emu: ZumaFullZ80Emulator, page: int, addr: int) -> None:
    emu.mem.physical[page * 0x4000 + (addr & 0x3FFF)] = 0xC9


def main() -> int:
    si = setting_index()
    bi = board_index()
    adv = expected_adventure(si, bi)
    gaunt = expected_gauntlet(si)

    assert [tuple(row) for row in db_rows("AdventureChainTable")] == adv
    assert db_rows("GauntletSettingTable") == gaunt
    assert (ROOT / "Build" / "level_chain_table.bin").read_bytes() == expected_chain_blob(adv, gaunt)
    assert len(adv) == 76
    assert len(gaunt) == 22 and all(len(row) == 4 for row in gaunt)
    assert adv[0] == (0, 0)
    assert adv[15] == (0, 15)   # 21-1 -> 1-2
    assert adv[33] == (0, 33)   # 21-2 -> 1-3
    assert adv[54] == (0, 54)   # 21-3 -> 1-4
    assert adv[-1] == (21, 75)
    assert gaunt[6][1] == si["lvl66"]  # Long Range Eagle — заполненный финальный Eagle-пресет

    runtime = (ASM / "level_runtime_table.inc").read_text(encoding="utf-8")
    assert "AdventureChainTable:" not in runtime
    assert "GauntletSettingTable:" not in runtime
    assert "LevelSelectApplyLevelClick:" not in runtime

    menu = (ASM / "MenuMain.asm").read_text(encoding="utf-8")
    assert "JP   FadeMenuToAdventure" in menu
    assert "JP   FadeMenuToLevelSelect" in menu

    main_asm = (ASM / "main.asm").read_text(encoding="utf-8")
    level_select = (ASM / "LevelSelect.asm").read_text(encoding="utf-8")
    spg = (ROOT / "spgbld_vdac2.ini").read_text(encoding="utf-8")
    assert "Block = #0000, #42, Build/level_chain_table.bin" in spg
    assert 'include "level_chain_table.inc"' not in main_asm
    assert "CALL OVL_ResolveCurrentModeSelection" in main_asm
    assert "CALL LevelSelectClampCurrent             ; Space/22-4 re-enters Gauntlet as last selectable 21-4" in menu
    assert "bonus 22-4 lose/retry -> 21-4" in main_asm
    assert "FadeGameplayToCurrentLevel:" in main_asm
    assert "JP   FadeGameplayToCurrentLevel" in main_asm
    assert "LEVEL_SELECT_COUNT     EQU LEVEL_RUNTIME_COUNT - 1" in main_asm
    assert "LevelSelectClampCurrent:" in level_select
    assert "LevelSelectIsLastSelectableLevel:" in level_select
    assert "JR   Z, .skip_next_control" in level_select
    assert "JP   LevelSelectDrawNextButton" in level_select
    assert "LevelSelectDrawNextButton:" in level_select
    assert "CALL LevelSelectIsLastSelectableLevel\n                JR   Z, .disabled\n                JP   LevelSelectDrawNext" in level_select
    assert "CP   LEVEL_SELECT_COUNT - 1" in level_select
    assert "CP   LEVEL_RUNTIME_COUNT" not in level_select

    emu = ZumaFullZ80Emulator(ROOT)
    sym = emu.sym
    emu.mem.pages[3] = 0x40

    emu.set_byte(sym["Core.AdventurePos"], 75)
    emu.set_byte(sym["Core.CurrentGameMode"], 0)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 21
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == 75
    assert emu.get_byte(sym["Core.CurrentDifficulty"]) == 3

    emu.set_byte(sym["Core.AdventurePos"], 99)
    emu.set_byte(sym["Core.CurrentGameMode"], 0)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.AdventurePos"]) == 0
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 0
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == 0

    emu.set_byte(sym["Core.CurrentLevel"], 6)
    emu.set_byte(sym["Core.CurrentDifficulty"], 1)
    emu.set_byte(sym["Core.CurrentGameMode"], 1)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == si["lvl66"]

    emu.set_byte(sym["Core.CurrentGameMode"], 0)
    emu.set_byte(sym["Core.AdventurePos"], 1)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 1
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == 1
    assert emu.get_byte(sym["Core.CurrentDifficulty"]) == 0

    for pos, diff, setting_id in ((15, 1, 15), (33, 2, 33), (54, 3, 54)):
        emu.set_byte(sym["Core.CurrentGameMode"], 0)
        emu.set_byte(sym["Core.AdventurePos"], pos)
        emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
        assert emu.get_byte(sym["Core.CurrentLevel"]) == 0
        assert emu.get_byte(sym["Core.CurrentDifficulty"]) == diff
        assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == setting_id

    emu.set_byte(sym["Core.CurrentGameMode"], 1)
    emu.set_byte(sym["Core.CurrentLevel"], 13)
    emu.set_byte(sym["Core.CurrentDifficulty"], 1)
    emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
    assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == si["lvl66"]

    for difficulty in (0, 1, 2, 3):
        emu.set_byte(sym["Core.CurrentGameMode"], 1)
        emu.set_byte(sym["Core.CurrentLevel"], 21)
        emu.set_byte(sym["Core.CurrentDifficulty"], difficulty)
        emu.call(sym["Core.OVL_ResolveCurrentModeSelection"])
        assert emu.get_byte(sym["Core.CurrentLevel"]) == 21
        assert emu.get_byte(sym["Core.CurrentDifficulty"]) == 3
        assert emu.get_byte(sym["Core.CurrentSettingIndex"]) == 75

    emu.mem.pages[3] = 0x41
    emu.set_byte(sym["Core.CurrentLevel"], 21)
    emu.set_byte(sym["Core.CurrentDifficulty"], 3)
    emu.call(sym["Core.LevelSelectClampCurrent"])
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 20
    assert emu.get_byte(sym["Core.CurrentDifficulty"]) == 3
    emu.set_byte(sym["Core.LevelSelectNextClick"], 1)
    emu.set_byte(sym["Core.LevelSelectBackClick"], 0)
    emu.call(sym["Core.LevelSelectApplyLevelClick"])
    assert emu.get_byte(sym["Core.CurrentLevel"]) == 20

    patch_ret(emu, 0x00, sym["UnpackAndUploadPage"])
    patch_ret(emu, 0x40, sym["Core.OVL_DrawNextLevelLoadingScreen"])
    patch_ret(emu, 0x05, sym["Core.LoadGameplayAssets"])
    patch_ret(emu, 0x05, sym["Core.DrawBlackLoadingFrame"])
    patch_ret(emu, 0x04, sym["Core.MainLoop"])

    def run_next(mode: int, level: int, difficulty: int, adventure_pos: int = 0):
        emu.set_byte(sym["Core.CurrentGameMode"], mode)
        emu.set_byte(sym["Core.CurrentLevel"], level)
        emu.set_byte(sym["Core.CurrentDifficulty"], difficulty)
        emu.set_byte(sym["Core.AdventurePos"], adventure_pos)
        emu.call(sym["Core.LoadNextLevelWithLoading"])
        return (
            emu.get_byte(sym["Core.CurrentLevel"]),
            emu.get_byte(sym["Core.CurrentDifficulty"]),
            emu.get_byte(sym["Core.AdventurePos"]),
            emu.get_byte(sym["Core.CurrentSettingIndex"]),
        )

    for difficulty in (0, 1, 2):
        assert run_next(1, 20, difficulty)[:2] == (0, difficulty + 1)
    assert run_next(1, 20, 3)[:2] == (21, 3)
    assert run_next(1, 21, 3)[:2] == (0, 0)  # win from bonus 22-4 wraps to 1-1

    for difficulty, expected_pos in ((0, 15), (1, 33), (2, 54)):
        assert run_next(0, 20, difficulty) == (0, difficulty + 1, expected_pos, expected_pos)
    assert run_next(0, 20, 3, 74) == (21, 3, 75, 75)
    assert run_next(0, 21, 3, 75) == (0, 0, 0, 0)

    print("PASS: Adventure/Gauntlet цепочки, L21 rank handoff, Space gate и level-select endpoints")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
